from pathlib import Path
import json
import openpyxl as xl
import re
import argparse
from pathlib import Path
import os
from openpyxl.worksheet.hyperlink import Hyperlink
import sys

from xlsxconvert import xlc

characters = []
codes = ['--Branch--','----','--Decision End--',]
infoFlag = True
commentFlag = False

bold = xl.styles.Font(b=True)
underline = xl.styles.Font(u='single')


actTablePath = Path('gamedata/excel/activity_table.json')
stgTablePath = Path('gamedata/excel/stage_table.json')
zoneTablePath = Path('gamedata/excel/zone_table.json')
mainStoryPath = Path('gamedata/story/obt/main')
reviewPath = Path('gamedata/excel/story_review_table.json')
metaPath = Path('gamedata/excel/story_review_meta_table.json')
infoPath = Path('gamedata/excel/storyinfo.json')


class Event():
    """
    Basic class of ak event
    """

    def __init__(self, data_dir:Path, remote_dir, lang:str, eventid:str):
        self.root_dir = data_dir/lang
        self.lang = lang
        self.eventid = eventid
        self.remote_dir = remote_dir+f'/{lang}'
        with open(data_dir/lang/reviewPath, encoding='utf-8') as reviewFile:
            self.review = json.load(reviewFile)[eventid]

        with open(data_dir/lang/infoPath, encoding='utf-8') as infoFile:
            self.info = json.load(infoFile)[eventid]

        self.entryType = self.review['entryType']
        self.name = self.review['name']
        self.storyList = self.review['infoUnlockDatas']


    def __getitem__(self, idx):
        return Story(self, self.storyList[idx])

    def __len__(self):
        return len(self.storyList)

class Story():
    """
    Class of indivisual story
    """
    def __init__(self, event:Event, storyData:dict):
        self.remote_dir = event.remote_dir
        self.root_dir = event.root_dir
        self.lang = event.lang
        self.eventid = event.eventid
        self.eventName = event.name
        self.entryType = event.entryType
        self.storyData = storyData
        self.storyCode = storyData['storyCode']
        self.avgTag = storyData['avgTag'] if storyData['avgTag'] else ''
        self.storyName = storyData['storyName']
        self.storyInfo = self.root_dir/'gamedata/story/[uc]{}.txt'.format(storyData['storyInfo'])
        self.storyTxt = self.root_dir/'gamedata/story/{}.txt'.format(storyData['storyTxt'])
        self.storyTxt_remote = self.remote_dir+f'/gamedata/story/{storyData["storyTxt"]}.json'
        self.f = storyData['storyTxt']
        self.storyInfo_remote = event.info[self.f]



async def web(eventid, lang):
    from pyodide.http import pyfetch

    event = Event(Path('ArknightsGameData'),'https://raw.githubusercontent.com/050644zf/ArknightsStoryJson/main' ,lang, eventid)

    filename = f'{event.eventid}_{event.name}.xlsx'

    txtList = []
    for story in event:
        txtList.append(story.storyTxt)

    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'Menu'
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 80

    ws.append([event.eventid, event.name])
    ws['B1'].font = bold
    ws.append([])
    ws.append(['storyCode', 'storyName', 'avgTag',
                'storyTxt.stem', 'storyInfo' if infoFlag else None])
    for cell in ws['A3':'E3'][0]:
        cell.font = bold

    for idx, story in enumerate(event):
        if infoFlag:
            storyInfo = story.storyInfo_remote
        else:
            storyInfo = None

        ws.append([story.storyCode, story.storyName,
                    story.avgTag, story.storyTxt.stem, storyInfo])
        loc = f"'{story.storyTxt.stem}'!A1"
        cell = ws.cell(idx+4, 4)
        cell.font = underline
        cell.hyperlink = Hyperlink(
            ref='', location=loc, display=story.storyTxt.stem)

        for i in range(1, 5):
            ws.cell(
                idx+4, i).alignment = xl.styles.Alignment(vertical='center')
        if ws.cell(idx+3, 1).value == story.storyCode and len(story.storyCode):
            ws.merge_cells(start_column=1, end_column=1,
                            start_row=idx+3, end_row=idx+4)
            ws.merge_cells(start_column=2, end_column=2,
                            start_row=idx+3, end_row=idx+4)

        for row in ws[ws.dimensions]:
            row[4].alignment = xl.styles.Alignment(wrap_text=True)


        print(f"{len(txtList)} txt files dectected")

        for (txtindex, txtFile) in enumerate(txtList):
            ws = wb.create_sheet(title=txtFile.stem)
            ws.column_dimensions['A'].hidden= True
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 70


            response = await pyfetch(story.storyTxt_remote)
            rawList = await response.json()
        
            xlc(ws,rawList)

            for row in ws[ws.dimensions]:
                for cell in row:
                    cell.alignment = xl.styles.Alignment(wrap_text=True)
            print(
                f"\rTxt file {txtFile.name} exported \033[1m({txtindex+1}/{len(txtList)})\033[m                 ")

        ws = wb.create_sheet(title='Characters')
        ws.append(['<Characters>'])
        for name in characters:
            ws.append([name])

        ws.append(['<Codes>'])
        for name in codes:
            ws.append([name])

        print("Character sheet exported")
        try:
            wb.save(filename=filename)
            print(f"\033[92mExported to \033[1m{filename}\033[m")
            return filename
        except PermissionError:
            print('\033[91mPermissionError: Fail to save the file, maybe because you have already opened the file! Close the file and rerun.\033[m')
            sys.exit('PermissionError')